from __future__ import annotations

import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import PENDING_REPORT_DIR


INVOICE_HEADERS = [
    "Barcade",
    "Cert No",
    "Cert",
    "Shape",
    "Size",
    "Color",
    "Clarity",
    "Fluo",
    "Cut",
    "Pol",
    "Sym",
    "L",
    "W",
    "D",
    "Depth %",
    "Table %",
    "RAP",
    "Discount",
    "$/ct",
    "Total $",
]


COLUMN_WIDTHS = {
    "A": 16,
    "B": 16,
    "C": 10,
    "D": 13,
    "E": 10,
    "F": 10,
    "G": 11,
    "H": 12,
    "I": 10,
    "J": 10,
    "K": 10,
    "L": 10,
    "M": 10,
    "N": 10,
    "O": 11,
    "P": 11,
    "Q": 12,
    "R": 13,
    "S": 13,
    "T": 14,
}


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TOTAL_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

THIN_BORDER = Border(
    left=Side(style="thin", color="B7B7B7"),
    right=Side(style="thin", color="B7B7B7"),
    top=Side(style="thin", color="B7B7B7"),
    bottom=Side(style="thin", color="B7B7B7"),
)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value),
    ).strip()


def clean_number(
    value: Any,
    default: float = 0.0,
) -> float:
    if value is None:
        return default

    if isinstance(value, (int, float)):
        return float(value)

    cleaned = re.sub(
        r"[^0-9.\-]",
        "",
        str(value),
    )

    if cleaned in {"", "-", ".", "-."}:
        return default

    try:
        return float(cleaned)
    except ValueError:
        return default


def normalize_header(value: Any) -> str:
    return re.sub(
        r"[^a-z0-9]",
        "",
        normalize_text(value).lower(),
    )


def build_portal_mapping(
    portal_headers: list[str],
    portal_values: list[Any],
) -> dict[str, Any]:
    mapping: dict[str, Any] = {}

    for header, value in zip(
        portal_headers,
        portal_values,
    ):
        mapping[
            normalize_header(header)
        ] = value

    return mapping


def get_portal_value(
    mapping: dict[str, Any],
    *possible_headers: str,
) -> Any:
    for header in possible_headers:
        key = normalize_header(header)

        if key in mapping:
            return mapping[key]

    return ""


def adjusted_discount(
    portal_discount: Any,
) -> float:
    """
    Increase the discount by one percentage point.

    Examples:
      -97.83 becomes -98.83
       97.83 becomes -98.83
    """

    value = clean_number(portal_discount)

    return -(abs(value) + 1.0)


def invoice_filename(
    processing_date: date | None = None,
) -> str:
    processing_date = (
        processing_date or date.today()
    )

    return (
        f"Keyzar Invoice "
        f"{processing_date.strftime('%m%d%y')}.xlsx"
    )


def invoice_path(
    processing_date: date | None = None,
) -> Path:
    PENDING_REPORT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    return (
        PENDING_REPORT_DIR
        / invoice_filename(processing_date)
    )


def get_current_invoice_path() -> Path | None:
    """
    Return the newest unsent Keyzar invoice.

    Normally there is only one pending invoice because Send Now deletes
    it after a successful Outlook submission. Using the newest file also
    handles a pending invoice that crosses midnight.
    """

    PENDING_REPORT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    candidates = [
        path
        for path in PENDING_REPORT_DIR.glob(
            "Keyzar Invoice *.xlsx"
        )
        if path.is_file()
    ]

    if not candidates:
        return None

    return max(
        candidates,
        key=lambda path: path.stat().st_mtime,
    )


def style_header(sheet) -> None:
    for column_index, header in enumerate(
        INVOICE_HEADERS,
        start=1,
    ):
        cell = sheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = THIN_BORDER

    sheet.row_dimensions[1].height = 24

    for column_letter, width in (
        COLUMN_WIDTHS.items()
    ):
        sheet.column_dimensions[
            column_letter
        ].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:T1"


def create_invoice_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Keyzar Invoice"

    style_header(sheet)

    workbook.save(path)


def find_vendor_row(
    sheet,
    vendor_id: str,
) -> int | None:
    for row_number in range(
        2,
        sheet.max_row + 1,
    ):
        existing_vendor = normalize_text(
            sheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if (
            existing_vendor.upper()
            == vendor_id.upper()
        ):
            return row_number

    return None


def remove_existing_totals(sheet) -> None:
    """
    Remove the previous totals row before appending
    another stone.
    """

    for row_number in range(
        sheet.max_row,
        1,
        -1,
    ):
        marker = normalize_text(
            sheet.cell(
                row=row_number,
                column=1,
            ).value
        ).upper()

        if marker == "TOTAL":
            sheet.delete_rows(
                row_number,
                1,
            )
            return


def write_totals_row(
    sheet,
    first_data_row: int,
    last_data_row: int,
) -> None:
    total_row = last_data_row + 1

    sheet.cell(
        row=total_row,
        column=1,
        value="TOTAL",
    )

    # Total carat under Size column.
    sheet.cell(
        row=total_row,
        column=5,
        value=(
            f"=SUM(E{first_data_row}:"
            f"E{last_data_row})"
        ),
    )

    # Grand total under Total $ column.
    sheet.cell(
        row=total_row,
        column=20,
        value=(
            f"=SUM(T{first_data_row}:"
            f"T{last_data_row})"
        ),
    )

    for column in range(1, 21):
        cell = sheet.cell(
            row=total_row,
            column=column,
        )

        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    sheet.cell(
        row=total_row,
        column=5,
    ).number_format = "0.00"

    sheet.cell(
        row=total_row,
        column=20,
    ).number_format = '$#,##0.00'

    sheet.cell(
        row=total_row,
        column=1,
    ).alignment = Alignment(
        horizontal="center",
    )


def append_stone_to_daily_report(
    *,
    order_date: date | datetime | None = None,
    order_number: str | None = None,
    vendor_id: str,
    portal_headers: list[str],
    portal_values: list[Any],
    processing_date: date | None = None,
) -> tuple[Path, bool]:
    """
    Append one successfully blocked stone.

    The invoice is grouped by the date the bot
    blocks/processes the stone, not the original
    forwarded email date.
    """

    del order_date
    del order_number

    processing_date = (
        processing_date or date.today()
    )

    path = invoice_path(processing_date)

    if not path.exists():
        create_invoice_workbook(path)

    workbook = load_workbook(path)
    sheet = workbook["Keyzar Invoice"]

    if find_vendor_row(sheet, vendor_id):
        print(
            f"[EXCEL] Vendor ID already exists: "
            f"{vendor_id}"
        )

        workbook.close()
        return path, False

    remove_existing_totals(sheet)

    mapping = build_portal_mapping(
        portal_headers,
        portal_values,
    )

    report_number = get_portal_value(
        mapping,
        "Report No",
        "Report Number",
        "Certificate No",
    )

    lab = get_portal_value(
        mapping,
        "Lab",
        "Cert",
    )

    shape = get_portal_value(
        mapping,
        "Shape",
    )

    carat = clean_number(
        get_portal_value(
            mapping,
            "Carat",
            "Size",
        )
    )

    color = get_portal_value(
        mapping,
        "Color",
    )

    clarity = get_portal_value(
        mapping,
        "Clarity",
    )

    fluorescence = get_portal_value(
        mapping,
        "Fluorescence",
        "Fluo",
    )

    cut = get_portal_value(
        mapping,
        "Cut",
    )

    polish = get_portal_value(
        mapping,
        "Polish",
        "Pol",
    )

    symmetry = get_portal_value(
        mapping,
        "Symmetry",
        "Sym",
    )

    length = clean_number(
        get_portal_value(
            mapping,
            "Length",
            "L",
        )
    )

    width = clean_number(
        get_portal_value(
            mapping,
            "Width",
            "W",
        )
    )

    depth = clean_number(
        get_portal_value(
            mapping,
            "Depth",
            "D",
        )
    )

    depth_percentage = clean_number(
        get_portal_value(
            mapping,
            "Depth %",
            "Depth Percent",
        )
    )

    table_percentage = clean_number(
        get_portal_value(
            mapping,
            "Table %",
            "Table Percent",
        )
    )

    rap = clean_number(
        get_portal_value(
            mapping,
            "Rap",
            "RAP",
        )
    )

    discount = adjusted_discount(
        get_portal_value(
            mapping,
            "Discount %",
            "Discount",
        )
    )

    row_number = sheet.max_row + 1

    values = [
        vendor_id,
        report_number,
        lab,
        shape,
        carat,
        color,
        clarity,
        fluorescence,
        cut,
        polish,
        symmetry,
        length,
        width,
        depth,
        depth_percentage,
        table_percentage,
        rap,
        discount,
        None,
        None,
    ]

    for column_number, value in enumerate(
        values,
        start=1,
    ):
        cell = sheet.cell(
            row=row_number,
            column=column_number,
            value=value,
        )

        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # $/ct = RAP + RAP × discount percentage.
    sheet.cell(
        row=row_number,
        column=19,
        value=f"=ROUND(Q{row_number}*(1+R{row_number}/100),2)",
    )

    # Total = $/ct × carat.
    sheet.cell(
        row=row_number,
        column=20,
        value=f"=ROUND(S{row_number}*E{row_number},2)",
    )

    for column in (
        5,
        12,
        13,
        14,
        15,
        16,
    ):
        sheet.cell(
            row=row_number,
            column=column,
        ).number_format = "0.00"

    sheet.cell(
        row=row_number,
        column=17,
    ).number_format = '$#,##0.00'

    sheet.cell(
        row=row_number,
        column=18,
    ).number_format = '0.00%'

    # Excel percentage values must be stored as decimal values.
    sheet.cell(
        row=row_number,
        column=18,
        value=discount / 100,
    )

    # Formula now refers directly to the decimal percentage.
    sheet.cell(
        row=row_number,
        column=19,
        value=f"=ROUND(Q{row_number}*(1+R{row_number}),2)",
    )

    sheet.cell(
        row=row_number,
        column=19,
    ).number_format = '$#,##0.00'

    sheet.cell(
        row=row_number,
        column=20,
    ).number_format = '$#,##0.00'

    write_totals_row(
        sheet=sheet,
        first_data_row=2,
        last_data_row=row_number,
    )

    workbook.save(path)
    workbook.close()

    print(
        f"[EXCEL] Added {vendor_id} to {path.name}"
    )

    return path, True